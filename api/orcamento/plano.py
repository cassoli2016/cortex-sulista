"""Importa o orçamento PLANEJADO (planilha da diretoria) como versão.

Diferença essencial para `derivacao.py`: lá o baseline é DEDUZIDO do histórico
(mês espelho × fator); aqui ele é o número que a diretoria planejou. Não se
regenera nem se recalcula — vem do arquivo.

Dois descasamentos de granularidade que este módulo resolve, e a razão de cada
escolha:

1. **A planilha é por AGRUPADOR; o orçamento é por CONTA.** O plano diz
   "CV - COMBUSTÍVEL: R$ 607 mil em janeiro", não diz quanto em cada uma das
   contas de combustível. O valor é rateado entre as contas do agrupador **na
   proporção do histórico** — o total do agrupador bate EXATO com o plano e o
   detalhe por conta continua existindo para o drill-down. O rateio é uma
   alocação, não um plano: fica marcado em `origem='plano_rateio'`.

2. **A receita do plano está toda numa linha só.** O arquivo põe os R$ 144 mi
   em `010.001 - FROTA/LOCADOS` e zero em agregados, terceiros e pedágio,
   enquanto a DRE realizada separa por modalidade. Importar linha a linha
   acusaria FROTA estourando o orçamento e AGREGADOS 100% abaixo — duas
   variações gigantes e ambas falsas. Por decisão do usuário a receita entra
   como **linha única no total**: as cinco linhas do plano viram um bucket só,
   rateado entre TODAS as contas de receita pela proporção do histórico. Assim
   o mix por modalidade sai realista e o total bate exato.

Linha de subtotal da planilha (`040.010 CUSTO FIXO`, `060.010 OVERHEAD`) é
ignorada de propósito: somá-la com os filhos duplicaria o valor.
"""
from __future__ import annotations

import re
import unicodedata
from pathlib import Path

from api import db
from api.orcamento import armazenamento as arm
from api.orcamento.sql import AGRUP_CONTA_SQL, HIST_CONTA_SQL


def _norm(s) -> str:
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.upper().split())


# A planilha e o ERP escrevem o mesmo agrupador de formas diferentes. Apelido
# explícito em vez de casamento aproximado: "parecido" erraria em silêncio, e
# um agrupador trocado move dinheiro de linha na DRE.
APELIDOS = {
    "CV - SINISTROS VEICULOS": "CV - SINISTROS",
    "RESULTADO VENDA IMOBILIZADO": "RECEITA - VENDA IMOBILIZADO",
    # A planilha escreve "TRUBUTARIOS". Sem este apelido, R$ 16,35 MILHOES de
    # credito tributario ficariam fora do orcamento em silencio - e o total
    # ainda fecharia contra o resultado do exercicio, porque a linha some
    # inteira dos dois lados. Foi pego so na reconciliacao.
    "CR - CREDITOS TRUBUTARIOS": "CR - CREDITOS",
    "CR - CREDITOS TRIBUTARIOS": "CR - CREDITOS",
}

# Prefixos de código que são RECEITA BRUTA na planilha. Entram num bucket só.
COD_RECEITA = ("010.",)

# Códigos que são SUBTOTAL (têm filhos na planilha) — nunca importados.
# `001` e `010` são os dois totais de receita; o resto são níveis intermediários.
COD_SUBTOTAL = {"001", "010", "020", "030", "040", "050", "060", "070", "080",
                "090", "100", "110", "040.010", "040.020", "040.030",
                "060.010", "060.020", "060.030", "080.010", "080.020"}

# Bucket sintético da receita. Não é agrupador do ERP: é a decisão de tratar a
# receita no total, e o nome precisa dizer isso em qualquer lugar que apareça.
BUCKET_RECEITA = "*RECEITA BRUTA (total do plano)"


def ler_planilha(caminho: str | Path) -> dict:
    """Lê o arquivo e devolve {'linhas': [...], 'meses': [1..12]}.

    Cada linha: codigo, nome, subtotal (bool), receita (bool), valores {mes: v}.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - dependência declarada
        raise RuntimeError("openpyxl não instalado no servidor.") from None

    wb = load_workbook(caminho, data_only=True)
    sh = wb[wb.sheetnames[0]]
    linhas = []
    for row in sh.iter_rows(values_only=True):
        rot = str(row[0] or "").strip()
        m = re.match(r"^([\d.]+)\s*-\s*(.+)$", rot)
        if not m:
            continue
        cod, nome = m.group(1), m.group(2).strip()
        valores = {}
        for mes in range(1, 13):
            v = row[mes] if mes < len(row) else None
            if isinstance(v, (int, float)):
                valores[mes] = float(v)
        linhas.append({
            "codigo": cod, "nome": nome,
            "subtotal": cod in COD_SUBTOTAL,
            "receita": cod.startswith(COD_RECEITA),
            "valores": valores,
            "total": round(sum(valores.values()), 2),
        })
    if not linhas:
        raise ValueError("Nenhuma linha de orçamento reconhecida na planilha.")
    return {"linhas": linhas, "arquivo": Path(caminho).name}


def _contas_por_agrupador() -> dict[str, list[str]]:
    fora: dict[str, list[str]] = {}
    for r in db.query(AGRUP_CONTA_SQL):
        fora.setdefault(_norm(r["agrupador"]), []).append(r["conta"])
    return fora


def carregar_historico(meses: list[str]) -> dict[str, float]:
    """Total em MÓDULO por conta no período. UMA consulta, não uma por
    agrupador — a primeira versão chamava isto dentro do laço e disparava ~40
    varreduras do razão contra o AVA, o que pendurava a importação.

    Módulo porque custo é negativo e receita positiva: com sinal, uma conta de
    estorno viraria peso negativo e inverteria o rateio.
    """
    de = f"{meses[0]}-01"
    ano, mes = int(meses[-1][:4]), int(meses[-1][5:7]) + 1
    if mes == 13:
        mes, ano = 1, ano + 1
    rows = db.query(HIST_CONTA_SQL, {"de": de, "ate": f"{ano:04d}-{mes:02d}-01"})
    fora: dict[str, float] = {}
    for r in rows:
        fora[r["conta"]] = fora.get(r["conta"], 0.0) + abs(float(r["valor"] or 0))
    return fora


def _peso_historico(contas: list[str], hist: dict[str, float]) -> dict[str, float]:
    """Peso de cada conta dentro do agrupador.

    Conta sem histórico nenhum recebe peso igual — o plano existe e o dinheiro
    precisa cair em algum lugar; jogar tudo numa conta só concentraria o
    orçamento inteiro num lançamento que pode ser marginal.
    """
    if not contas:
        return {}
    bruto = {c: hist.get(c, 0.0) for c in contas}
    soma = sum(bruto.values())
    if soma <= 0:
        return {c: 1.0 / len(contas) for c in contas}
    return {c: v / soma for c, v in bruto.items()}


def _ratear(valor: float, pesos: dict[str, float]) -> dict[str, float]:
    """Distribui `valor` pelos pesos, jogando o resíduo de centavos na MAIOR
    fatia. Sem isso a soma das contas erra o plano por alguns centavos por mês
    — e doze meses de resíduo vira uma diferença que ninguém explica."""
    if not pesos:
        return {}
    itens = sorted(pesos.items(), key=lambda kv: -kv[1])
    fora, acum = {}, 0.0
    for c, p in itens[1:]:
        v = round(valor * p, 2)
        fora[c] = v
        acum += v
    fora[itens[0][0]] = round(valor - acum, 2)
    return fora


def preparar(caminho: str | Path, meses_hist: list[str]) -> dict:
    """Casa a planilha com os agrupadores do ERP e rateia por conta.

    Não grava nada — devolve o que seria gravado mais o relatório do que não
    casou, para a decisão acontecer ANTES de o número entrar no sistema.
    """
    lido = ler_planilha(caminho)
    por_ag = _contas_por_agrupador()
    hist = carregar_historico(meses_hist)

    # Receita: as cinco linhas do plano viram um bucket só (decisão do usuário
    # — o plano não rateia por modalidade e comparar linha a linha produziria
    # variação falsa dos dois lados).
    receita_mes: dict[int, float] = {}
    alvos: dict[str, dict[int, float]] = {}
    nao_casaram: list[dict] = []
    ignorados: list[dict] = []

    for l in lido["linhas"]:
        if l["subtotal"]:
            ignorados.append({"codigo": l["codigo"], "nome": l["nome"],
                              "total": l["total"], "motivo": "subtotal da planilha"})
            continue
        if l["receita"]:
            for m, v in l["valores"].items():
                receita_mes[m] = receita_mes.get(m, 0.0) + v
            continue
        chave = _norm(APELIDOS.get(_norm(l["nome"]), l["nome"]))
        if chave not in por_ag:
            nao_casaram.append({"codigo": l["codigo"], "nome": l["nome"],
                                "total": l["total"]})
            continue
        destino = alvos.setdefault(chave, {})
        for m, v in l["valores"].items():
            destino[m] = destino.get(m, 0.0) + v

    # contas de receita = união das contas de todos os agrupadores de receita
    contas_receita = sorted({c for ag, cs in por_ag.items()
                             if ag.startswith("RECEITA") for c in cs})

    linhas_orc: list[dict] = []
    conferencia: list[dict] = []

    def _emitir(rotulo: str, contas: list[str], por_mes: dict[int, float],
                origem: str):
        pesos = _peso_historico(contas, hist)
        total_plano = round(sum(por_mes.values()), 2)
        total_rateado = 0.0
        for mes, valor in sorted(por_mes.items()):
            for conta, v in _ratear(valor, pesos).items():
                linhas_orc.append({"conta": conta, "mes": mes,
                                   "valor_baseline": v, "origem": origem,
                                   "meses_com_dado": 12})
                total_rateado += v
        conferencia.append({"agrupador": rotulo, "contas": len(contas),
                            "plano": total_plano,
                            "rateado": round(total_rateado, 2),
                            "diferenca": round(total_rateado - total_plano, 2)})

    if receita_mes:
        _emitir(BUCKET_RECEITA, contas_receita, receita_mes, "plano_receita_total")
    for chave, por_mes in alvos.items():
        _emitir(chave, por_ag[chave], por_mes, "plano_rateio")

    return {
        "arquivo": lido["arquivo"],
        "linhas": linhas_orc,
        "conferencia": conferencia,
        "nao_casaram": nao_casaram,
        "ignorados": ignorados,
        "receita_total": round(sum(receita_mes.values()), 2),
        "valor_importado": round(sum(l["valor_baseline"] for l in linhas_orc), 2),
        "valor_nao_importado": round(sum(x["total"] for x in nao_casaram), 2),
        "contas": len({l["conta"] for l in linhas_orc}),
    }


def importar(caminho: str | Path, ano: int, rotulo: str, quem: str,
             meses_hist: list[str], path=None) -> dict:
    """Cria uma versão NOVA com o plano. Nunca toca nas versões existentes —
    a derivada do histórico continua lá para comparação."""
    prep = preparar(caminho, meses_hist)
    caminho_db = path or arm.DB_PATH
    arm.init_db(caminho_db)
    versao_id = arm.criar_versao(caminho_db, ano, rotulo, 0.0, quem,
                                 metodo="plano", meses_base=[])
    arm.gravar_baseline(caminho_db, versao_id, prep["linhas"])
    return {**prep, "versao_id": versao_id, "ano": ano, "rotulo": rotulo}
