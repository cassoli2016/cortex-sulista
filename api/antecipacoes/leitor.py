"""Lê o arquivo do portal e devolve os títulos no modelo canônico.

Formatos aceitos: .xls (BIFF, via xlrd), .xlsx (via openpyxl, se instalado) e
.csv/.txt. O primeiro arquivo recebido é BIFF de verdade — não é HTML com
extensão .xls, que é o disfarce mais comum de portal web e por isso está
tratado explicitamente abaixo.

Três coisas que este módulo faz e que não são óbvias:

1. **Acha o cabeçalho.** Portal costuma exportar com linhas de título/filtro
   antes da tabela. Procura nas primeiras linhas a que melhor pontua contra
   algum modelo, em vez de assumir linha 0.
2. **Descarta a linha de TOTAL.** O arquivo da Maxion termina com uma linha
   só com o total na coluna Nominal (929.085,9999999992). Somada como título
   dobraria o valor; ignorada em silêncio perderia a única conferência que o
   arquivo oferece. Ela vira `total_declarado` e é reconciliada.
3. **Reconcilia.** Se a soma dos títulos não bate com o total declarado, o
   resultado sai com `divergencia` preenchida. Importar calado um arquivo que
   não fecha é o jeito mais rápido de pôr número errado na tela.
"""
from __future__ import annotations

import csv
import io
from datetime import date
from pathlib import Path

from api.antecipacoes import modelos as md
from api.antecipacoes import valores as vl

# Quantas linhas do começo podem ser cabeçalho/preâmbulo.
LINHAS_PROCURA_CABECALHO = 12
# Tolerância da reconciliação: o total do arquivo vem com resíduo de ponto
# flutuante (929085,9999999992 para 929086,00). Um centavo por título é ruído
# de arredondamento; acima disso é título faltando.
TOLERANCIA = 0.05


class ArquivoInvalido(Exception):
    """Erro que a TELA mostra. A mensagem tem de dizer o que fazer."""


def _celulas_xls(dados: bytes) -> list[list]:
    try:
        import xlrd
    except ImportError:  # pragma: no cover - dependência declarada
        raise ArquivoInvalido("Suporte a .xls indisponível no servidor.") from None
    try:
        wb = xlrd.open_workbook(file_contents=dados)
    except Exception as exc:  # noqa: BLE001
        raise ArquivoInvalido(
            "Não foi possível abrir a planilha. Confira se o arquivo é um "
            "Excel válido e não está protegido por senha.") from exc
    sh = wb.sheet_by_index(0)
    # xlrd entrega data como serial + o modo do arquivo; converter aqui evita
    # que cada modelo precise saber do datemode.
    linhas = []
    for r in range(sh.nrows):
        linha = []
        for c in range(sh.ncols):
            v = sh.cell_value(r, c)
            if sh.cell_type(r, c) == 3:      # XL_CELL_DATE
                import xlrd.xldate
                try:
                    v = xlrd.xldate.xldate_as_datetime(v, wb.datemode).date()
                except Exception:  # noqa: BLE001
                    pass
            linha.append(v)
        linhas.append(linha)
    return linhas


def _celulas_xlsx(dados: bytes) -> list[list]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ArquivoInvalido(
            "Este servidor lê .xls e .csv. Salve a planilha como .xls ou "
            "exporte em CSV.") from None
    wb = load_workbook(io.BytesIO(dados), read_only=True, data_only=True)
    return [list(l) for l in wb[wb.sheetnames[0]].iter_rows(values_only=True)]


def _celulas_csv(dados: bytes) -> list[list]:
    # Portal brasileiro exporta em latin-1 com frequência; utf-8 primeiro
    # porque é o que quebra de forma detectável.
    for cod in ("utf-8-sig", "latin-1"):
        try:
            texto = dados.decode(cod)
            break
        except UnicodeDecodeError:
            continue
    else:  # pragma: no cover - latin-1 aceita qualquer byte
        raise ArquivoInvalido("Não foi possível ler o texto do arquivo.")
    amostra = texto[:4000]
    try:
        dial = csv.Sniffer().sniff(amostra, delimiters=";,\t|")
    except csv.Error:
        dial = csv.excel
        dial.delimiter = ";"          # padrão do Excel em pt-BR
    return [l for l in csv.reader(io.StringIO(texto), dial)]


def celulas(nome: str, dados: bytes) -> list[list]:
    """Matriz de células a partir do arquivo bruto."""
    if not dados:
        raise ArquivoInvalido("Arquivo vazio.")
    ext = Path(nome or "").suffix.lower()
    cabeca = dados[:8]

    # Assinatura ganha da extensão: portal que exporta HTML com nome .xls é
    # comum, e a mensagem de erro do xlrd nesse caso não ajuda ninguém.
    if cabeca.startswith(b"\xd0\xcf\x11\xe0"):          # OLE2 = .xls de verdade
        return _celulas_xls(dados)
    if cabeca.startswith(b"PK\x03\x04"):                 # zip = .xlsx
        return _celulas_xlsx(dados)
    inicio = dados[:400].lstrip().lower()
    if inicio.startswith(b"<") and (b"<table" in dados[:4000].lower()
                                    or b"<html" in inicio):
        raise ArquivoInvalido(
            "Este arquivo é uma página HTML com extensão .xls (alguns portais "
            "exportam assim). Abra no Excel e salve como .xls ou .csv.")
    if ext in (".csv", ".txt"):
        return _celulas_csv(dados)
    raise ArquivoInvalido(
        "Formato não reconhecido. Envie a planilha do portal em .xls, .xlsx "
        "ou .csv.")


def _achar_cabecalho(linhas: list[list]) -> tuple[int, md.Modelo, float]:
    melhor = (-1, None, 0.0)
    for i, linha in enumerate(linhas[:LINHAS_PROCURA_CABECALHO]):
        modelo, nota = md.escolher([vl.texto(c) for c in linha])
        if modelo and nota > melhor[2]:
            melhor = (i, modelo, nota)
    if melhor[1] is None:
        raise ArquivoInvalido(
            "Não reconheci o layout desta planilha. Os portais já conhecidos "
            "são: " + ", ".join(m.rotulo for m in md.MODELOS) + ". Se for um "
            "portal novo, envie o arquivo para cadastrarmos o modelo.")
    return melhor


def _linha_de_total(vals: dict, mapa: dict) -> bool:
    """Linha de rodapé: sem identificação, mas com valor.

    Genérico de propósito — todo portal fecha a planilha com um total e cada
    um o põe numa coluna diferente. O que define é a AUSÊNCIA de título e de
    documento junto com a presença de valor.
    """
    tem_id = bool(vals.get("titulo") or vals.get("documento")
                  or vals.get("id_portal") or vals.get("chave"))
    return not tem_id and (vals.get("valor_nominal") or vals.get("valor_saldo"))


def ler(nome: str, dados: bytes) -> dict:
    """Lê o arquivo e devolve títulos canônicos + reconciliação.

    Nunca levanta para valor de célula ruim: a linha problemática vai para
    `rejeitadas` com o motivo. Um título com data ilegível não pode impedir
    a importação dos outros 225 — mas some da tela se ninguém contar.
    """
    linhas = celulas(nome, dados)
    i_cab, modelo, nota = _achar_cabecalho(linhas)
    cabecalho = [vl.texto(c) for c in linhas[i_cab]]
    mapa = modelo.mapear(cabecalho)

    def _campo(linha, campo):
        i = mapa.get(campo)
        return linha[i] if (i is not None and i < len(linha)) else None

    titulos, rejeitadas = [], []
    total_declarado = None

    for n, linha in enumerate(linhas[i_cab + 1:], start=i_cab + 2):
        if not any(str(c).strip() for c in linha if c is not None):
            continue
        crus = {c: _campo(linha, c) for c in mapa}
        vals = {
            "titulo": vl.texto(crus.get("titulo")),
            "documento": vl.texto(crus.get("documento")),
            "id_portal": vl.texto(crus.get("id_portal")),
            "chave": vl.texto(crus.get("chave")),
            "valor_nominal": vl.numero(crus.get("valor_nominal")),
            "valor_saldo": vl.numero(crus.get("valor_saldo")),
        }
        if _linha_de_total(vals, mapa):
            total_declarado = vals["valor_nominal"] or vals["valor_saldo"]
            continue

        venc = vl.data(crus.get("vencimento"))
        nominal = vals["valor_nominal"]
        if nominal is None:
            rejeitadas.append({"linha": n, "motivo": "valor nominal ilegível",
                               "conteudo": vl.texto(crus.get("valor_nominal"))})
            continue
        if venc is None:
            rejeitadas.append({"linha": n, "motivo": "vencimento ilegível",
                               "conteudo": vl.texto(crus.get("vencimento"))})
            continue

        # Saldo ausente = título íntegro. O portal só preenche quando houve
        # pagamento parcial; assumir zero zeraria a antecipação inteira.
        saldo = vals["valor_saldo"]
        titulos.append({
            "titulo": vals["titulo"],
            "documento": vals["documento"],
            "emissao": vl.data(crus.get("emissao")),
            "vencimento": venc,
            "valor_nominal": round(nominal, 2),
            "valor_saldo": round(saldo if saldo is not None else nominal, 2),
            "antecipavel": vl.booleano(crus.get("antecipavel")),
            "situacao": vl.texto(crus.get("situacao")),
            "cnpj_cedente": vl.cnpj(crus.get("cnpj_cedente")),
            "nome_cedente": vl.texto(crus.get("nome_cedente")),
            "cnpj_sacado": vl.cnpj(crus.get("cnpj_sacado")),
            "nome_sacado": vl.texto(crus.get("nome_sacado")),
            "chave": vals["chave"],
            "id_portal": vals["id_portal"],
            "linha": n,
        })

    if not titulos:
        raise ArquivoInvalido(
            "O layout foi reconhecido, mas nenhuma linha de título pôde ser "
            "lida. Confira se a exportação trouxe dados.")

    soma = round(sum(t["valor_nominal"] for t in titulos), 2)
    divergencia = None
    if total_declarado is not None:
        d = round(soma - total_declarado, 2)
        if abs(d) > TOLERANCIA:
            divergencia = d

    return {
        "portal": modelo.nome,
        "portal_rotulo": modelo.rotulo,
        "confianca": round(nota, 3),
        "arquivo": nome,
        "linha_cabecalho": i_cab + 1,
        "colunas_lidas": sorted(mapa),
        "colunas_ignoradas": [c for i, c in enumerate(cabecalho)
                              if c and i not in mapa.values()],
        "titulos": titulos,
        "rejeitadas": rejeitadas,
        "total_declarado": total_declarado,
        "total_calculado": soma,
        "divergencia": divergencia,
    }


def resumir(lido: dict) -> dict:
    """Números que a tela mostra antes de o usuário confirmar a importação."""
    ts = lido["titulos"]
    antec = [t for t in ts if t["antecipavel"] is not False]
    hoje = date.today()
    sacados: dict = {}
    cedentes: dict = {}
    for t in ts:
        s = sacados.setdefault(t["cnpj_sacado"],
                               {"cnpj": t["cnpj_sacado"], "nome": t["nome_sacado"],
                                "titulos": 0, "valor": 0.0})
        s["titulos"] += 1
        s["valor"] = round(s["valor"] + t["valor_saldo"], 2)
        # O cedente identifica a FILIAL: o arquivo da Maxion tem títulos de
        # dois CNPJs nossos (0001-23 e 0020-96). Somar tudo junto esconderia
        # de qual filial é o recebível.
        c = cedentes.setdefault(t["cnpj_cedente"],
                                {"cnpj": t["cnpj_cedente"], "nome": t["nome_cedente"],
                                 "titulos": 0, "valor": 0.0})
        c["titulos"] += 1
        c["valor"] = round(c["valor"] + t["valor_saldo"], 2)

    vencs = [t["vencimento"] for t in ts]
    prazos = [(v - hoje).days for v in vencs]
    return {
        "titulos": len(ts),
        "valor_nominal": round(sum(t["valor_nominal"] for t in ts), 2),
        "valor_saldo": round(sum(t["valor_saldo"] for t in ts), 2),
        "antecipaveis": len(antec),
        "valor_antecipavel": round(sum(t["valor_saldo"] for t in antec), 2),
        "vencimento_de": min(vencs).isoformat(),
        "vencimento_ate": max(vencs).isoformat(),
        # prazo médio ponderado pelo valor: é o que define o custo do deságio.
        # A média simples deixaria 200 títulos pequenos mandarem no número.
        "prazo_medio_dias": round(
            sum((t["vencimento"] - hoje).days * t["valor_saldo"] for t in ts)
            / max(0.01, sum(t["valor_saldo"] for t in ts))),
        "prazo_min_dias": min(prazos),
        "prazo_max_dias": max(prazos),
        "vencidos": sum(1 for p in prazos if p < 0),
        "sacados": sorted(sacados.values(), key=lambda x: -x["valor"]),
        "cedentes": sorted(cedentes.values(), key=lambda x: -x["valor"]),
    }
